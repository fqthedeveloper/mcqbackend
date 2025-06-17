import csv
import io
import json
from tkinter.font import Font
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from .models import *
from .serializers import *
from .utils import process_excel
from django.utils import timezone
from django.http import Http404, HttpResponse
import pandas as pd
from django.db.models import Q
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny
from rest_framework.authtoken.models import Token
from django.db import IntegrityError
from django.db import IntegrityError, transaction
from rest_framework.permissions import IsAuthenticated
from .permissions import IsAdminUserOnly
from rest_framework.authentication import TokenAuthentication
from rest_framework.exceptions import PermissionDenied, ValidationError as DRFValidationError
from django.conf import settings
import random
import re
from channels.generic.websocket import AsyncWebsocketConsumer
import subprocess
from asgiref.sync import sync_to_async

User = get_user_model()

logger = logging.getLogger(__name__)
User = get_user_model()

class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer


class LoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        username_or_email = request.data.get('username_or_email')
        password = request.data.get('password')

        if not username_or_email or not password:
            return Response({"error": "Username/email and password required."}, status=status.HTTP_400_BAD_REQUEST)

        user_qs = User.objects.filter(Q(username=username_or_email) | Q(email=username_or_email))
        if not user_qs.exists():
            return Response({"username_or_email": ["User with this email or username does not exist."]}, status=status.HTTP_400_BAD_REQUEST)

        # Try authenticate by username first
        user = authenticate(username=username_or_email, password=password)

        # If not found, try by email username
        if not user:
            try:
                email_user = User.objects.get(email=username_or_email)
                user = authenticate(username=email_user.username, password=password)
            except User.DoesNotExist:
                user = None

        if not user:
            return Response({"password": ["Incorrect password."]}, status=status.HTTP_400_BAD_REQUEST)

        try:
            token, created = Token.objects.get_or_create(user=user)
        except Exception:
            return Response({"error": "Authentication system error. Please contact support."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "role": getattr(user, "user_type", None),
            "token": token.key,
            "refresh_token": token.key,  # Placeholder; replace with real refresh token system
            "force_password_change": getattr(user, "force_password_change", False),
            "first_name": user.first_name,
            "is_verified": getattr(user, "is_verified", False)
        }, status=status.HTTP_200_OK)



class StudentViewSet(viewsets.ModelViewSet):
    queryset = User.objects.filter(user_type='student')
    serializer_class = StudentCreateSerializer
    permission_classes = [IsAuthenticated, IsAdminUserOnly]
    authentication_classes = [TokenAuthentication]

    def list(self, request, *args, **kwargs):
        print("Logged in user:", request.user)
        print("Returning students:", self.get_queryset())
        return super().list(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            self.perform_create(serializer)
        except IntegrityError:
            return Response(
                {"detail": "A user with this email or username already exists."},
                status=status.HTTP_400_BAD_REQUEST
            )

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        old_email = instance.email

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        new_email = serializer.validated_data.get('email', old_email)

        if new_email != old_email:
            send_mail(
                subject='Your email has been updated',
                message=(
                    f'Hello {instance.first_name},\n\n'
                    f'Your email address has been changed to {new_email}.\n'
                    'If you did not request this change, please contact support immediately.'
                ),
                from_email='no-reply@example.com',
                recipient_list=[new_email],
                fail_silently=False,
            )

        return Response(serializer.data, status=status.HTTP_200_OK)

class ForcePasswordChangeView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [TokenAuthentication]

    def post(self, request):
        serializer = PasswordChangeSerializer(data=request.data)
        if serializer.is_valid():
            request.user.set_password(serializer.validated_data['new_password'])
            request.user.force_password_change = False
            request.user.save()
            return Response({"message": "Password changed successfully."})
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class SubjectViewSet(viewsets.ModelViewSet):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    


class QuestionViewSet(viewsets.ModelViewSet):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer

    @action(detail=False, methods=['GET'])
    def download_format(self, request):
        data = {
            'Subject': ['Mathematics', 'Science'],
            'Question': ['What is 2+2?', 'Water boils at?'],
            'Option A': ['3', '90°C'],
            'Option B': ['4', '100°C'],
            'Option C': ['5', '110°C'],
            'Option D': ['6', '120°C'],
            'Correct Answers': ['B', 'B'],
            'Marks': [1, 1],
            'Is Multi': [False, False]
        }
        df = pd.DataFrame(data)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="question_format.xlsx"'
        df.to_excel(response, index=False)
        return response

    @action(detail=False, methods=['POST'])
    def bulk_upload(self, request):
        # Ensure file exists in request.FILES
        if 'file' not in request.FILES:
            return Response(
                {'status': 'error', 'message': 'No file uploaded'}, 
                status=400
            )
            
        file = request.FILES['file']
        
        try:
            # Handle different Excel formats
            if file.name.endswith('.xlsx'):
                df = pd.read_excel(file, engine='openpyxl')
            elif file.name.endswith('.xls'):
                df = pd.read_excel(file, engine='xlrd')
            else:
                return Response(
                    {'status': 'error', 'message': 'Invalid file format'},
                    status=400
                )
        except Exception as e:
            return Response(
                {'status': 'error', 'message': f'Failed to read file: {str(e)}'},
                status=400
            )

        required_columns = [
            'Subject', 'Question',
            'Option A', 'Option B', 'Option C', 'Option D',
            'Correct Answers', 'Marks', 'Is Multi'
        ]
        missing_cols = [col for col in required_columns if col not in df.columns]
        if missing_cols:
            return Response({'status': 'error', 'message': f'Missing columns: {", ".join(missing_cols)}'}, status=400)

        try:
            with transaction.atomic():
                for idx, row in df.iterrows():
                    raw_subject = row['Subject']
                    if pd.isna(raw_subject) or str(raw_subject).strip() == '':
                        raise ValueError(f"Row {idx+2}: 'Subject' cannot be blank.")

                    subject_obj, _ = Subject.objects.get_or_create(name=str(raw_subject).strip())

                    correct_answers = str(row['Correct Answers']).replace(' ', '').upper()
                    is_multi_val = row['Is Multi']
                    is_multi = str(is_multi_val).strip().lower() in ['true', '1', 'yes']

                    if not is_multi:
                        if ',' in correct_answers:
                            raise ValueError(f"Row {idx+2}: multiple answers '{correct_answers}' given for a single-answer question.")
                        if correct_answers not in ['A', 'B', 'C', 'D']:
                            raise ValueError(f"Row {idx+2}: invalid single correct answer '{correct_answers}'.")
                    else:
                        for opt in correct_answers.split(','):
                            if opt not in ['A', 'B', 'C', 'D']:
                                raise ValueError(f"Row {idx+2}: invalid correct option '{opt}'.")

                    Question.objects.create(
                        subject=subject_obj,
                        text=str(row['Question']).strip(),
                        option_a=str(row['Option A']).strip(),
                        option_b=str(row['Option B']).strip(),
                        option_c=str(row['Option C']).strip(),
                        option_d=str(row['Option D']).strip(),
                        correct_option=correct_answers,
                        marks=int(row['Marks']),
                        is_multi=is_multi
                    )

                return Response({'status': 'success', 'message': 'Questions uploaded successfully'})
        except ValueError as ve:
            return Response({'status': 'error', 'message': str(ve)}, status=400)
        except Exception as e:
            return Response({'status': 'error', 'message': str(e)}, status=400)


class ExamViewSet(viewsets.ModelViewSet):
    queryset = Exam.objects.all()
    serializer_class = ExamSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.user_type == 'student':
            return Exam.objects.filter(is_published=True)
        return Exam.objects.all()

    @action(detail=True, methods=['POST'])
    def publish(self, request, pk=None):
        exam = self.get_object()
        exam.is_published = True
        exam.save()
        message = request.data.get('message', f'New exam "{exam.title}" scheduled')
        students = User.objects.filter(user_type='student', is_active=True)
        for student in students:
            if student.email:
                send_mail(
                    subject=f"New Exam: {exam.title}",
                    message=message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[student.email],
                    fail_silently=True,
                )
        return Response({'status': 'published', 'recipients': students.count()})

    @action(detail=True, methods=['POST'])
    def unpublish(self, request, pk=None):
        exam = self.get_object()
        exam.is_published = False
        exam.save()
        return Response({'status': 'unpublished'})
    

class ExamSessionViewSet(viewsets.ModelViewSet):
    serializer_class = ExamSessionSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [TokenAuthentication]

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser:
            return ExamSession.objects.all()
        elif user.user_type == 'student':
            return ExamSession.objects.filter(student=user)
        elif user.user_type == 'teacher':
            return ExamSession.objects.filter(exam__subject__teacher=user)
        return ExamSession.objects.none()

    def get_object(self):
        try:
            return super().get_object()
        except Http404:
            user = self.request.user
            session_id = self.kwargs['pk']
            exam_id = self.request.query_params.get('exam_id')
            
            if not exam_id:
                logger.warning("Missing exam_id for session creation")
                raise DRFValidationError(
                    {'error': 'exam_id parameter is required'},
                    code=status.HTTP_400_BAD_REQUEST
                )
            
            try:
                exam = Exam.objects.get(id=exam_id)
                if exam.mode == 'practical':
                    return self.create_practical_session(user, exam, session_id)
                else:
                    logger.warning("Invalid exam mode for session creation")
                    raise DRFValidationError(
                        {'error': 'Only practical exams can be created this way'},
                        code=status.HTTP_400_BAD_REQUEST
                    )
            except Exam.DoesNotExist:
                logger.warning("Exam not found")
                raise DRFValidationError(
                    {'error': 'Exam not found'},
                    code=status.HTTP_404_NOT_FOUND
                )

    def create_practical_session(self, user, exam, session_id=None):
        if exam.mode == 'strict':
            if ExamSession.objects.filter(student=user, exam=exam, is_completed=True).exists():
                raise DRFValidationError(
                    {'error': 'Strict exam already completed'},
                    code=status.HTTP_400_BAD_REQUEST
                )

        try:
            existing_session = ExamSession.objects.get(id=session_id)
            if existing_session.student != user:
                raise PermissionDenied("Session already exists for another user")
            return existing_session
        except ExamSession.DoesNotExist:
            pass

        session = ExamSession(
            id=session_id,
            student=user,
            exam=exam,
            is_completed=False
        )

        try:
            session.full_clean()
            session.save()
            return session
        except Exception as e:
            raise DRFValidationError(
                {'error': str(e)},
                code=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['post'], url_path='validate-exam')
    def validate_exam_session(self, request):
        student = request.user
        exam_id = request.data.get('exam')
        
        if not exam_id:
            return Response(
                {'error': 'Exam ID is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            exam = Exam.objects.get(id=exam_id)
        except Exam.DoesNotExist:
            return Response(
                {'error': 'Exam not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Check for existing active session
        existing_session = ExamSession.objects.filter(
            student=student, 
            exam=exam, 
            is_completed=False
        ).first()
        
        if existing_session:
            serializer = self.get_serializer(existing_session)
            return Response(serializer.data)

        # Create new session
        session = ExamSession(student=student, exam=exam, is_completed=False)
        try:
            session.full_clean()
            session.save()
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = self.get_serializer(session)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        student = request.user
        exam_id = request.data.get('exam')
        if not exam_id:
            return Response({'error': 'Exam required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            exam = Exam.objects.get(id=exam_id)
        except Exam.DoesNotExist:
            return Response({'error': 'Exam not found'}, status=status.HTTP_404_NOT_FOUND)

        # Check for existing active session
        existing_session = ExamSession.objects.filter(
            student=student, 
            exam=exam, 
            is_completed=False
        ).first()
        
        if existing_session:
            serializer = self.get_serializer(existing_session)
            return Response(serializer.data, status=status.HTTP_200_OK)

        # Create new session
        session = ExamSession(student=student, exam=exam, is_completed=False)
        try:
            session.full_clean()
            session.save()
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        serializer = self.get_serializer(session)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['POST'])
    def start_exam(self, request, pk=None):
        try:
            session = self.get_object()
            if session.student != request.user:
                raise PermissionDenied("Invalid session owner")
            if not session.start_time:
                session.start_time = timezone.now()
                session.save()
            return Response({'status': 'exam started'})
        except ExamSession.DoesNotExist:
            return Response(
                {'error': 'Session not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )

    @action(detail=True, methods=['POST'], url_path='save_progress')
    def save_progress(self, request, pk=None):
        session = self.get_object()
        user = request.user

        if session.student != user:
            raise PermissionDenied("Invalid session owner")

        if session.is_completed:
            return Response(
                {'error': 'Exam already submitted'},
                status=status.HTTP_400_BAD_REQUEST
            )

        answers_data = request.data.get('answers', [])
        elapsed_time = request.data.get('elapsed_time', 0)

        if not isinstance(answers_data, list):
            return Response(
                {'error': 'Answers must be a list'}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():
            # Delete existing answers only if new ones are provided
            if answers_data:
                Answer.objects.filter(session=session).delete()

            for answer_data in answers_data:
                question_id = answer_data.get('question')
                selected_answers = answer_data.get('selected_answers', '')
                if not question_id:
                    continue
                Answer.objects.create(
                    session=session,
                    question_id=question_id,
                    selected_answers=selected_answers
                )

            session.elapsed_time = elapsed_time
            session.save()
            
        return Response({'status': 'progress saved'}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['POST'], url_path='save_terminal')
    def save_terminal(self, request, pk=None):
        session = self.get_object()
        if session.student != request.user:
            raise PermissionDenied("Invalid session owner")
        
        if session.is_completed:
            return Response(
                {'error': 'Exam already submitted'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        terminal_output = request.data.get('terminal_output', '')
        session.terminal_output = terminal_output
        session.save()
        
        return Response({'status': 'terminal output saved'})

    @action(detail=True, methods=['POST'])
    def submit_exam(self, request, pk=None):
        session = self.get_object()
        user = request.user

        if session.student != user:
            raise PermissionDenied("Invalid session owner")
        if session.is_completed:
            return Response(
                {'error': 'Exam already submitted'}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        if session.exam.mode == 'practical':
            return self._submit_practical_exam(session, request)
        else:
            return self._submit_mcq_exam(session, request)

    def _submit_practical_exam(self, session, request):
        terminal_output = request.data.get('terminal_output', '')
        session.terminal_output = terminal_output
        
        answers = PracticalAnswer.objects.filter(session=session).select_related('task')
        total_marks = 0
        score = 0
        details = {}
        
        with transaction.atomic():
            for answer in answers:
                task = answer.task
                total_marks += task.marks
                
                try:
                    pattern = re.compile(task.expected_output)
                    is_verified = bool(pattern.search(answer.output))
                except re.error:
                    is_verified = False
                    
                earned = task.marks if is_verified else 0
                score += earned
                
                answer.is_verified = is_verified
                answer.save()
                
                details[task.id] = {
                    'command_used': answer.command_used,
                    'output': answer.output,
                    'expected_output': task.expected_output,
                    'is_verified': is_verified,
                    'marks': task.marks,
                    'earned': earned
                }
            
            result = Result.objects.create(
                session=session,
                score=score,
                total_marks=total_marks,
                details=details
            )
            
            session.end_time = timezone.now()
            session.is_completed = True
            session.save()
        
        return Response(ResultSerializer(result).data, status=status.HTTP_200_OK)

    def _submit_mcq_exam(self, session, request):
        answers_data = request.data.get('answers', [])
        elapsed_time = request.data.get('elapsed_time', 0)
        termination_reason = request.data.get('termination_reason', None)

        if not isinstance(answers_data, list):
            return Response(
                {'error': 'Answers must be a list'}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():
            # Delete existing answers only if new ones are provided
            if answers_data:
                Answer.objects.filter(session=session).delete()

            for answer_data in answers_data:
                question_id = answer_data.get('question')
                selected_answers = answer_data.get('selected_answers', '')
                if not question_id:
                    continue
                Answer.objects.create(
                    session=session,
                    question_id=question_id,
                    selected_answers=selected_answers
                )

            answers = Answer.objects.filter(session=session).select_related('question')
            total_marks = 0
            score = 0
            details = {}

            for answer in answers:
                question = answer.question
                total_marks += question.marks
                correct_set = set(question.correct_option.split(','))
                selected_set = set(answer.selected_answers.split(','))
                is_correct = correct_set == selected_set
                earned = question.marks if is_correct else 0
                if is_correct:
                    score += earned
                details[question.id] = {
                    'correct': list(correct_set),
                    'selected': list(selected_set),
                    'is_correct': is_correct,
                    'marks': question.marks,
                    'earned': earned
                }

            result = Result.objects.create(
                session=session,
                score=score,
                total_marks=total_marks,
                details=details
            )

            session.end_time = timezone.now()
            session.is_completed = True
            session.elapsed_time = elapsed_time
            if termination_reason:
                session.termination_reason = termination_reason
            session.save()

        return Response(ResultSerializer(result).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['GET'], url_path='download_answers')
    def download_answers(self, request, pk=None):
        session = self.get_object()
        
        if session.student != request.user and request.user.user_type != 'teacher':
            return Response(
                {'error': 'Permission denied'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        if session.exam.mode == 'practical':
            answers = PracticalAnswer.objects.filter(session=session).select_related('task')
        else:
            answers = Answer.objects.filter(session=session).select_related('question')
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{session.exam.title}_{session.student.username}_answers.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Exam Answers"
        
        if session.exam.mode == 'practical':
            headers = [
                'Task Title', 'Description', 'Command Used', 
                'Output', 'Expected Output', 'Marks', 'Status'
            ]
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
            
            for row_num, answer in enumerate(answers, 2):
                worksheet.cell(row=row_num, column=1, value=answer.task.title)
                worksheet.cell(row=row_num, column=2, value=answer.task.description)
                worksheet.cell(row=row_num, column=3, value=answer.command_used)
                worksheet.cell(row=row_num, column=4, value=answer.output)
                worksheet.cell(row=row_num, column=5, value=answer.task.expected_output)
                worksheet.cell(row=row_num, column=6, value=answer.task.marks)
                worksheet.cell(row=row_num, column=7, value='Verified' if answer.is_verified else 'Pending')
        else:
            headers = [
                'Question', 'Selected Answers', 'Correct Answers', 
                'Status', 'Marks', 'Earned'
            ]
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
            
            for row_num, answer in enumerate(answers, 2):
                question = answer.question
                correct_set = set(question.correct_option.split(','))
                selected_set = set(answer.selected_answers.split(','))
                is_correct = correct_set == selected_set
                
                worksheet.cell(row=row_num, column=1, value=question.text)
                worksheet.cell(row=row_num, column=2, value=answer.selected_answers)
                worksheet.cell(row=row_num, column=3, value=question.correct_option)
                worksheet.cell(row=row_num, column=4, value='Correct' if is_correct else 'Incorrect')
                worksheet.cell(row=row_num, column=5, value=question.marks)
                worksheet.cell(row=row_num, column=6, value=question.marks if is_correct else 0)
        
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
        
        workbook.save(response)
        return response
    
    
class AnswerViewSet(viewsets.ModelViewSet):
    serializer_class = AnswerSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [TokenAuthentication]

    def get_queryset(self):
        return Answer.objects.filter(session__student=self.request.user)

    def perform_create(self, serializer):
        session = serializer.validated_data.get('session')
        if session.student != self.request.user:
            raise PermissionDenied("Invalid session owner")
        if session.is_completed:
            raise ValidationError("Session already completed")
        serializer.save()



class ResultViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = ResultSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [TokenAuthentication]

    def get_queryset(self):
        user = self.request.user
        if user.user_type == 'student':
            return Result.objects.filter(session__student=user)
        return Result.objects.all()

    def get(self, request, session):
        result = get_object_or_404(Result, session=session, user=request.user)
        serializer = ResultSerializer(result, context={'request': request})
        return Response(serializer.data)

    def list(self, request, *args, **kwargs):
        user = request.user
        queryset = self.get_queryset().select_related('session__exam', 'session__student')
        

        result_list = []
        for result in queryset:
            session = result.session
            exam = session.exam
            student = session.student

            # Right & wrong answers count
            answers_qs = Answer.objects.filter(session=session).select_related('question')
            right = 0
            wrong = 0
            for answer in answers_qs:
                q = answer.question
                correct = q.correct_option.split(',') if ',' in q.correct_option else [q.correct_option]
                selected = answer.selected_answers.split(',') if answer.selected_answers else []
                if sorted(correct) == sorted(selected):
                    right += 1
                else:
                    wrong += 1

            percentage = (result.score / result.total_marks) * 100 if result.total_marks else 0
            pass_fail = "Pass" if percentage >= 80 else "Fail"

            result_list.append({
                'id': result.id,
                'session': session.id,
                'exam_title': exam.title,
                'student_name': student.get_full_name() or student.username,
                'date': session.end_time.isoformat() if session.end_time else None,
                'score': result.score,
                'total_marks': result.total_marks,
                'right_answers': right,
                'wrong_answers': wrong,
                'pass_fail': pass_fail
            })

        return Response(result_list)

    @action(detail=False, methods=['GET'], url_path='session/(?P<session_id>[0-9]+)')
    def by_session(self, request, session_id=None):
        session = get_object_or_404(ExamSession, id=session_id)
        user = request.user

        if user.user_type == 'student' and session.student != user:
            return Response({'error': 'Permission denied'}, status=403)

        result = get_object_or_404(Result, session=session)
        exam = session.exam
        answers_qs = Answer.objects.filter(session=session).select_related('question')

        total_right = 0
        total_wrong = 0
        detailed_answers = {}

        for answer in answers_qs:
            q = answer.question
            correct = q.correct_option.split(',') if ',' in q.correct_option else [q.correct_option]
            selected = answer.selected_answers.split(',') if answer.selected_answers else []
            is_correct = sorted(correct) == sorted(selected)

            if is_correct:
                total_right += 1
            else:
                total_wrong += 1

            detailed_answers[q.id] = {
                'question_text': q.text,
                'correct': correct,
                'selected': selected,
                'is_correct': is_correct,
                'marks': q.marks,
                'earned': q.marks if is_correct else 0,
                'explanation': getattr(q, 'explanation', '')
            }

        pass_status = "Pass" if result.score >= (0.8 * result.total_marks) else "Fail"

        response_data = {
            'exam_title': exam.title,
            'submitted_at': session.end_time,
            'duration': exam.duration,
            'student_name': session.student.get_full_name() or session.student.username,
            'score': result.score,
            'total_marks': result.total_marks,
            'right_answers': total_right,
            'wrong_answers': total_wrong,
            'result': pass_status,
            'details': detailed_answers,
        }

        return Response(response_data)
    

def generate_otp():
    return str(random.randint(100000, 999999))

class SendOTPView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        otp = generate_otp()

        EmailOTP.objects.update_or_create(user=user, defaults={'otp': otp})
        send_mail(
                'Your IRT MCQ APP OTP Code',
                f'Your OTP is {otp}',
                'noreply@example.com',
                [user.email],
                fail_silently=False,  # fix the typo here
            )
        return Response({'message': 'OTP sent successfully'})

class VerifyOTPView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        otp = request.data.get('otp')
        user = request.user

        try:
            email_otp = EmailOTP.objects.get(user=user)
            if email_otp.otp == otp:
                user.is_verified = True
                user.save()
                email_otp.delete()
                return Response({'message': 'Email verified successfully'})
            return Response({'error': 'Invalid OTP'}, status=400)
        except EmailOTP.DoesNotExist:
            return Response({'error': 'OTP not sent'}, status=400)

class GetStudentEmailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, user_id):
        try:
            student = User.objects.get(id=user_id, role='student')
            return Response({'email': student.email})
        except User.DoesNotExist:
            return Response({'error': 'Student not found'}, status=404)



class PracticalTaskViewSet(viewsets.ModelViewSet):
    queryset = PracticalTask.objects.all()
    serializer_class = PracticalTaskSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['GET'])
    def download_format(self, request):
        data = {
            'Title': ['Task 1', 'Task 2'],
            'Description': ['First task description', 'Second task description'],
            'Command Template': ['ls -la', 'grep "error" log.txt'],
            'Expected Output': ['.*file1.txt.*', '.*critical error.*'],
            'Marks': [10, 15]
        }
        df = pd.DataFrame(data)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="practical_task_format.xlsx"'
        df.to_excel(response, index=False)
        return response

    @action(detail=False, methods=['POST'])
    def bulk_upload(self, request):
        file = request.FILES.get('file')
        exam_id = request.data.get('exam_id')
        
        if not file:
            return Response(
                {'error': 'No file uploaded'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            if file.name.endswith('.xlsx'):
                df = pd.read_excel(file)
            elif file.name.endswith('.xls'):
                df = pd.read_excel(file, engine='xlrd')
            elif file.name.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                return Response(
                    {'error': 'Unsupported file format. Use .xlsx, .xls or .csv'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Exception as e:
            return Response(
                {'error': f'File processing error: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        required_columns = ['Title', 'Description', 'Command Template', 'Expected Output', 'Marks']
        missing_cols = [col for col in required_columns if col not in df.columns]
        if missing_cols:
            return Response(
                {'error': f'Missing columns: {", ".join(missing_cols)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        exam = None
        if exam_id:
            try:
                exam = Exam.objects.get(id=exam_id, mode='practical')
            except Exam.DoesNotExist:
                return Response(
                    {'error': 'Exam not found or not in practical mode'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        created_tasks = []
        errors = []
        
        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    task_data = {
                        'title': str(row['Title']).strip(),
                        'description': str(row['Description']).strip(),
                        'command_template': str(row['Command Template']).strip(),
                        'expected_output': str(row['Expected Output']).strip(),
                        'marks': int(row['Marks'])
                    }
                    
                    if not task_data['title']:
                        raise ValidationError('Title is required')
                        
                    if not task_data['expected_output']:
                        raise ValidationError('Expected output is required')
                    
                    # Validate regex pattern
                    try:
                        re.compile(task_data['expected_output'])
                    except re.error as e:
                        raise ValidationError(f'Invalid regex pattern: {str(e)}')
                    
                    # Create or update task
                    task, created = PracticalTask.objects.update_or_create(
                        title=task_data['title'],
                        defaults=task_data
                    )
                    created_tasks.append(task)
                    
                    # Link to exam if provided
                    if exam:
                        if not ExamPracticalTask.objects.filter(exam=exam, task=task).exists():
                            max_order = ExamPracticalTask.objects.filter(exam=exam).aggregate(
                                max_order=models.Max('order')
                            )['max_order'] or 0
                            
                            ExamPracticalTask.objects.create(
                                exam=exam,
                                task=task,
                                order=max_order + 1
                            )
                except Exception as e:
                    errors.append(f"Row {index+2}: {str(e)}")
        
        if errors:
            return Response(
                {'error': f'{len(errors)} errors occurred', 'details': errors},
                status=status.HTTP_207_MULTI_STATUS
            )
        
        return Response(
            {'success': f'Processed {len(created_tasks)} tasks (created/updated)'},
            status=status.HTTP_201_CREATED
        )