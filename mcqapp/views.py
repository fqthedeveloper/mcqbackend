from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.conf import settings
from django.core.mail import send_mail
from django.contrib.auth import authenticate, get_user_model
from django.db import IntegrityError, transaction
from rest_framework.filters import SearchFilter
import openpyxl
from rest_framework import viewsets, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.authentication import TokenAuthentication
from rest_framework.authtoken.models import Token
from rest_framework.exceptions import PermissionDenied
from .utils import generate_otp, send_otp_email
from rest_framework.permissions import IsAdminUser
from .models import (
    User, Subject, StudentSubjectEnrollment,
    Question, Exam, ExamSession, Answer, Result, EmailOTP
)
from .serializers import *
from .permissions import IsAdminUserOnly, IsStudentUserOnly
from mcqapp.models import Subject, PracticeQuestion, PracticeRun
from mcqapp.practice_service import (
    start_practice,
    submit_practice_answer,
    finish_practice,
)
from .totp import (
    generate_totp_secret,
    get_totp_uri,
    generate_qr_code_base64,
    verify_totp
)



UserModel = get_user_model()

# ======================================================
# USER
# ======================================================

class UserViewSet(viewsets.ModelViewSet):
    queryset = UserModel.objects.all()
    serializer_class = UserSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated, IsAdminUserOnly]


class StudentViewSet(viewsets.ModelViewSet):
    queryset = User.objects.filter(user_type='student')
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated, IsAdminUserOnly]

    def get_serializer_class(self):
        if self.action == 'create':
            return StudentCreateSerializer
        if self.action in ['update', 'partial_update']:
            return StudentUpdateSerializer
        return UserSerializer

    @action(detail=True, methods=['GET'])
    def enrolled_subjects(self, request, pk=None):
        enrollments = StudentSubjectEnrollment.objects.filter(
            student_id=pk, is_active=True
        )
        subjects = [e.subject for e in enrollments]
        return Response(SubjectSerializer(subjects, many=True).data)
    
    
# ======================================================
# AUTH
# ======================================================

class LoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        user = authenticate(
            username=request.data.get('username_or_email'),
            password=request.data.get('password')
        )

        if not user:
            return Response({'error': 'Invalid credentials'}, status=400)

        token, _ = Token.objects.get_or_create(user=user)

        return Response({
            'token': token.key,
            'full_name': user.first_name + ' ' + user.last_name,
            'role': user.user_type,
            'email': user.email,
            'force_password_change': user.force_password_change,
            'is_verified': user.is_verified,
        })

        
class SendEmailOTPView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get('email')

        if not email:
            return Response({'error': 'Email is required'}, status=400)

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=404)

        otp = generate_otp()

        EmailOTP.objects.update_or_create(
            user=user,
            defaults={'otp': otp, 'created_at': timezone.now()}
        )

        send_otp_email(user.email, otp)

        return Response({
            'message': 'OTP sent to email',
            'email': user.email
        })


class VerifyEmailOTPView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get('email')
        otp = request.data.get('otp')

        if not email or not otp:
            return Response({'error': 'Email and OTP are required'}, status=400)

        try:
            user = User.objects.get(email=email)
            email_otp = user.email_otp
        except (User.DoesNotExist, EmailOTP.DoesNotExist):
            return Response({'error': 'Invalid request'}, status=400)

        if email_otp.is_expired():
            return Response({'error': 'OTP expired'}, status=400)

        if email_otp.otp != otp:
            return Response({'error': 'Invalid OTP'}, status=400)

        # ✅ OTP VERIFIED
        user.is_verified = True
        user.save()
        email_otp.delete()

        # ✅ NO MFA → ISSUE TOKEN
        token, _ = Token.objects.get_or_create(user=user)

        return Response({
            'token': token.key,
            'full_name': f"{user.first_name} {user.last_name}",
            'role': user.user_type,
            'email': user.email,
            'force_password_change': user.force_password_change,
            'is_verified': user.is_verified,
            'login_type': 'otp'
        })



class ForcePasswordChangeView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [TokenAuthentication]

    def post(self, request):
        serializer = PasswordChangeSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        request.user.set_password(serializer.validated_data['new_password'])
        request.user.force_password_change = False
        request.user.save()

        return Response({'message': 'Password changed'})


class ForgotPasswordView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = ForgotPasswordSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        email = serializer.validated_data["email"]

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return Response({"message": "If email exists, reset link sent"})

        # ❌ ADMIN BLOCK
        if user.user_type == "admin":
            return Response(
                {"error": "Admin password reset not allowed"},
                status=403
            )

        token_obj = PasswordResetToken.objects.create(user=user)

        reset_link = f"http://localhost:3000/reset-password/{token_obj.token}"

        send_mail(
            subject="Reset Your Password",
            message=f"Click the link to reset your password:\n\n{reset_link}\n\nLink valid for 30 minutes.",
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[email],
            fail_silently=False,
        )

        return Response({"message": "Password reset link sent"})


class ResetPasswordView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = ResetPasswordSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        token = serializer.validated_data["token"]
        new_password = serializer.validated_data["new_password"]

        try:
            token_obj = PasswordResetToken.objects.get(token=token, is_used=False)
        except PasswordResetToken.DoesNotExist:
            return Response({"error": "Invalid reset link"}, status=400)

        if token_obj.is_expired():
            token_obj.delete()
            return Response({"error": "Reset link expired"}, status=400)

        user = token_obj.user

        if user.user_type == "admin":
            return Response(
                {"error": "Admin password reset not allowed"},
                status=403
            )

        user.set_password(new_password)
        user.force_password_change = False
        user.save()

        token_obj.is_used = True
        token_obj.save()

        return Response({"message": "Password reset successful"})
    
# ======================================================
# OTP
# ======================================================

class SendOTPView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = SendOTPSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        email = serializer.validated_data["email"]

        if email != request.user.email:
            return Response(
                {"error": "Email mismatch"},
                status=status.HTTP_400_BAD_REQUEST
            )

        otp = generate_otp()

        EmailOTP.objects.update_or_create(
            user=request.user,
            defaults={"otp": otp}
        )

        send_mail(
            subject="IRT MCQ Email Verification OTP",
            message=f"Your IRT MCQ Email Verification OTP is {otp}",
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[email],
            fail_silently=False
        )

        return Response(
            {"success": True, "message": "OTP sent"},
            status=status.HTTP_200_OK
        )


class VerifyOTPView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = VerifyOTPSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        email = serializer.validated_data["email"]
        otp = serializer.validated_data["otp"].strip()

        if email != request.user.email:
            return Response(
                {"error": "Email mismatch"},
                status=status.HTTP_400_BAD_REQUEST
            )

        record = EmailOTP.objects.filter(user=request.user).first()

        if not record:
            return Response(
                {"error": "OTP not found"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if record.is_expired():
            record.delete()
            return Response(
                {"error": "OTP expired"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if record.otp.strip() != otp:
            return Response(
                {"error": "Invalid OTP"},
                status=status.HTTP_400_BAD_REQUEST
            )

        request.user.is_verified = True
        request.user.save(update_fields=["is_verified"])

        record.delete()

        return Response(
            {"success": True, "message": "OTP verified"},
            status=status.HTTP_200_OK
        )

# ======================================================
# SUBJECT
# ======================================================

class SubjectViewSet(viewsets.ModelViewSet):
    serializer_class = SubjectSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    
    queryset = Subject.objects.all().order_by("name")

    def get_queryset(self):
        if self.request.user.user_type == 'student':
            return Subject.objects.filter(
                studentsubjectenrollment__student=self.request.user,
                studentsubjectenrollment__is_active=True
            )
        return Subject.objects.all()

# ======================================================
# QUESTION
# ======================================================

class QuestionViewSet(viewsets.ModelViewSet):
    serializer_class = QuestionSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    filter_backends = [SearchFilter]
    search_fields = ["text"]

    def get_queryset(self):
        qs = Question.objects.all()
        subject_id = self.request.query_params.get("subject")
        if subject_id:
            qs = qs.filter(subject_id=subject_id)
        return qs.order_by("-id")

    # ================= BULK UPLOAD =================
    @action(detail=False, methods=["POST"], url_path="upload-excel")
    def upload_excel(self, request):
        file = request.FILES.get("file")
        subject_id = request.data.get("subject")

        if not file or not subject_id:
            return Response(
                {"error": "file and subject are required"}, status=400
            )

        wb = openpyxl.load_workbook(file)
        sheet = wb.active
        created = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            (
                text,
                option_a,
                option_b,
                option_c,
                option_d,
                correct_option,
                marks,
                explanation,
            ) = row

            if not text:
                continue

            Question.objects.update_or_create(
                subject_id=subject_id,
                text=text,
                defaults={
                    "option_a": option_a,
                    "option_b": option_b,
                    "option_c": option_c,
                    "option_d": option_d,
                    "correct_option": correct_option,
                    "marks": marks or 1,
                    "explanation": explanation or "",
                },
            )
            created += 1

        return Response({"status": "success", "created": created})

    # ================= EXCEL TEMPLATE DOWNLOAD =================
    @action(detail=False, methods=["GET"], url_path="download-template")
    def download_template(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([
            "text",
            "option_a",
            "option_b",
            "option_c",
            "option_d",
            "correct_option",
            "marks",
            "explanation",
        ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="mcq_template.xlsx"'
        wb.save(response)
        return response

# ======================================================
# EXAM (MCQ ONLY)
# ======================================================

class ExamViewSet(viewsets.ModelViewSet):
    serializer_class = ExamSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user

        qs = (
            Exam.objects
            .select_related("subject")
            .prefetch_related("questions")
            .order_by("-created_at")
        )

        # STUDENT: published + enrolled subjects only
        if user.user_type == "student":
            enrolled_subjects = Subject.objects.filter(
                studentsubjectenrollment__student=user,
                studentsubjectenrollment__is_active=True
            )
            qs = qs.filter(
                is_published=True,
                subject__in=enrolled_subjects
            )

        subject_id = self.request.query_params.get("subject")
        if subject_id:
            qs = qs.filter(subject_id=subject_id)

        return qs


    @action(detail=True, methods=["post"])
    def publish(self, request, pk=None):
        exam = self.get_object()
        exam.is_published = True
        exam.save()
        return Response({"published": True})

    @action(detail=True, methods=["post"])
    def unpublish(self, request, pk=None):
        exam = self.get_object()
        exam.is_published = False
        exam.save()
        return Response({"published": False})
    
# ======================================================
# EXAM SESSION
# ======================================================
    

class ExamSessionViewSet(viewsets.ModelViewSet):
    serializer_class = ExamSessionSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.user_type == "student":
            return ExamSession.objects.filter(student=user)
        return ExamSession.objects.all()

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(
            data=request.data,
            context={"request": request},
        )
        serializer.is_valid(raise_exception=True)
        session = serializer.save()
        return Response(self.get_serializer(session).data, status=201)

    # 🔒 ABSOLUTELY SAFE SUBMIT
    @action(detail=True, methods=["post"])
    @transaction.atomic
    def submit(self, request, pk=None):
        session = (
            ExamSession.objects
            .select_for_update()
            .select_related("exam")
            .get(pk=pk)
        )

        if session.student != request.user:
            raise PermissionDenied("Not allowed")

        # 🔥 HARD STOP — NO SECOND SUBMIT
        if session.is_completed:
            return Response(
                {"detail": "Exam already submitted"},
                status=400
            )

        answers = request.data.get("answers", [])
        terminate_reason = request.data.get("terminate_reason", "manual")

        if terminate_reason not in dict(ExamSession.TERMINATE_CHOICES):
            terminate_reason = "manual"

        # 🔥 LOCK EXAM FIRST (IMPORTANT)
        session.is_completed = True
        session.end_time = timezone.now()
        session.terminate_reason = terminate_reason
        session.save(update_fields=[
            "is_completed",
            "end_time",
            "terminate_reason"
        ])

        # 🔥 DELETE OLD ANSWERS (SAFETY)
        Answer.objects.filter(session=session).delete()

        score = 0
        total = 0

        for a in answers:
            try:
                q = Question.objects.get(id=a["question"])
            except Question.DoesNotExist:
                continue

            total += q.marks

            if a["selected_answers"] == q.correct_option:
                score += q.marks

            Answer.objects.create(
                session=session,
                question=q,
                selected_answers=a["selected_answers"],
            )

        Result.objects.create(
            session=session,
            score=score,
            total_marks=total,
        )

        return Response({
            "score": score,
            "total": total,
            "terminate_reason": terminate_reason,
        })

# ======================================================
# ANSWER
# ======================================================

class AnswerViewSet(viewsets.ModelViewSet):
    queryset = Answer.objects.all()
    serializer_class = AnswerSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]


# ======================================================
# RESULT
# ======================================================

class ResultViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = ResultSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user

        # ✅ STUDENT → only own results
        if user.user_type == "student":
            return Result.objects.filter(session__student=user)

        # ✅ ADMIN → all students
        return Result.objects.select_related(
            "session__student",
            "session__exam"
        )

    # ✅ DETAIL BY SESSION (ADMIN + STUDENT SAFE)
    @action(
        detail=False,
        methods=["get"],
        url_path=r"session/(?P<session_id>[^/.]+)"
    )
    def by_session(self, request, session_id=None):
        user = request.user

        if user.user_type == "student":
            result = get_object_or_404(
                Result,
                session__id=session_id,
                session__student=user
            )
        else:
            result = get_object_or_404(
                Result,
                session__id=session_id
            )

        serializer = self.get_serializer(result)
        return Response(serializer.data)


# ======================================================
# DASHBOARDS
# ======================================================

class StudentDashboardView(APIView):
    permission_classes = [IsAuthenticated, IsStudentUserOnly]
    authentication_classes = [TokenAuthentication]

    def get(self, request):
        return Response({
            'subjects': SubjectSerializer(
                Subject.objects.filter(
                    studentsubjectenrollment__student=request.user,
                    studentsubjectenrollment__is_active=True
                ),
                many=True
            ).data,
            'exams': ExamSerializer(
                Exam.objects.filter(
                    subject__studentsubjectenrollment__student=request.user,
                    subject__studentsubjectenrollment__is_active=True,
                    is_published=True
                ),
                many=True
            ).data,
            'active_sessions': ExamSessionSerializer(
                ExamSession.objects.filter(student=request.user, is_completed=False),
                many=True
            ).data
        })


class AdminDashboardView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUserOnly]
    authentication_classes = [TokenAuthentication]

    def get(self, request):
        return Response({
            'total_students': UserModel.objects.filter(user_type='student').count(),
            'total_subjects': Subject.objects.count(),
            'total_questions': Question.objects.count(),
            'total_exams': Exam.objects.count(),
            'active_exams': Exam.objects.filter(is_published=True).count(),
            'total_results': Result.objects.count(),
        })


class MyProfileView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user

        # IMPORTANT: decide verification source
        # If you use Django default email verification
        # adjust this logic if needed
        is_verified = getattr(user, "is_verified", False)

        return Response({
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "is_verified": is_verified,
        }, status=status.HTTP_200_OK)

    def put(self, request):
        user = request.user

        user.first_name = request.data.get("first_name", user.first_name)
        user.last_name = request.data.get("last_name", user.last_name)
        user.email = request.data.get("email", user.email)

        user.save()

        return Response({
            "username": user.username,
            "email": user.email,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "is_verified": getattr(user, "is_verified", False),
        }, status=status.HTTP_200_OK)
        
        
class PracticeQuestionMapView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request):
        subject_id = request.data.get("subject_id")
        difficulty = request.data.get("difficulty")
        question_ids = request.data.get("question_ids", [])

        if not subject_id or not difficulty or not question_ids:
            return Response(
                {"detail": "subject_id, difficulty, question_ids required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ❗ already mapped questions (ANY difficulty)
        already_mapped = set(
            PracticeQuestion.objects.filter(
                subject_id=subject_id
            ).values_list("question_id", flat=True)
        )

        created = 0
        skipped = 0

        for qid in question_ids:
            if qid in already_mapped:
                skipped += 1
                continue

            PracticeQuestion.objects.create(
                subject_id=subject_id,
                question_id=qid,
                difficulty=difficulty
            )
            created += 1

        return Response({
            "status": "success",
            "mapped": created,
            "skipped": skipped
        })


# ===============================
# PRACTICE STATS
# ===============================
class PracticeStatsView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        subject_id = request.query_params.get("subject_id")

        if not subject_id:
            return Response({"easy": 0, "medium": 0, "hard": 0})

        qs = PracticeQuestion.objects.filter(subject_id=subject_id)

        return Response({
            "easy": qs.filter(difficulty="easy").count(),
            "medium": qs.filter(difficulty="medium").count(),
            "hard": qs.filter(difficulty="hard").count(),
        })
        
        

class StartPractice(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        subject_id = request.data.get("subject_id")
        difficulty = request.data.get("difficulty")

        if not subject_id or not difficulty:
            return Response(
                {"detail": "subject_id and difficulty are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        subject = get_object_or_404(Subject, id=int(subject_id))

        run, result = start_practice(request.user, subject, difficulty)

        if run is None:
            return Response(
                {"detail": result},
                status=status.HTTP_400_BAD_REQUEST
            )

        return Response({
            "run_id": run.id,
            "duration": run.duration_minutes,
            "questions": PracticeQuestionSerializer(result, many=True).data
        })


class SubmitPracticeAnswer(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        run_id = request.data.get("run_id")
        pq_id = request.data.get("practice_question_id")

        if not run_id or not pq_id:
            return Response({"detail": "Invalid data"}, status=400)

        run = get_object_or_404(PracticeRun, id=run_id)
        pq = get_object_or_404(PracticeQuestion, id=pq_id)

        submit_practice_answer(run, pq, request.data.get("selected_answers"))
        return Response({"status": "saved"})


class FinishPractice(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        run_id = request.data.get("run_id")

        if not run_id:
            return Response({"detail": "run_id required"}, status=400)

        run = get_object_or_404(PracticeRun, id=run_id)
        return Response(finish_practice(run))
    
    
class EnableTOTPView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user

        if user.is_totp_enabled:
            return Response({"error": "TOTP already enabled"}, status=400)

        secret = generate_totp_secret()
        uri = get_totp_uri(user, secret)
        qr_code = generate_qr_code_base64(uri)

        user.totp_secret = secret
        user.save()

        return Response({
            "qr_code": qr_code,
            "secret": secret
        })


class ConfirmTOTPView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        code = request.data.get("code")

        if not user.totp_secret:
            return Response({"error": "TOTP not initialized"}, status=400)

        if not verify_totp(user.totp_secret, code):
            return Response({"error": "Invalid TOTP"}, status=400)

        user.is_totp_enabled = True
        user.totp_enabled_at = timezone.now()
        user.save()

        return Response({"message": "TOTP enabled successfully"})


class VerifyLoginTOTPView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        code = request.data.get("code")

        if not user.is_totp_enabled:
            return Response({"error": "TOTP not enabled"}, status=400)

        if not verify_totp(user.totp_secret, code):
            return Response({"error": "Invalid TOTP"}, status=400)

        return Response({"message": "TOTP verified"})


class AdminResetTOTPView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request, user_id):
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({"error": "User not found"}, status=404)

        user.is_totp_enabled = False
        user.totp_secret = None
        user.totp_enabled_at = None
        user.save()

        return Response({
            "message": f"TOTP reset successfully for {user.email}"
        })